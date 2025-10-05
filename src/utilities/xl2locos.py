import openpyxl
import os
import json

fn = os.path.join(os.getcwd(), "data", "locos.xlsx")
wb = openpyxl.load_workbook(filename = fn)
ws = wb.active

row = 2  # skip title row

jloco = {}
while True:
    loco = ws.cell(row=row, column=1).value
    if loco is None:
        break
    
    desc = ws.cell(row=row, column=2).value
    start = ws.cell(row=row, column=3).value
    slow = ws.cell(row=row, column=4).value
    medium = ws.cell(row=row, column=5).value
    fast = ws.cell(row=row, column=6).value
    acc = ws.cell(row=row, column=7).value
    dec = ws.cell(row=row, column=8).value
    
    jloco[loco] = {"desc": desc, "prof": {"start": start, "slow": slow, "medium": medium, "fast": fast, "acc": acc, "dec": dec}}
    
    row += 1
    
jfn = os.path.join(os.getcwd(), "data", "newlocos.json")
with open(jfn, "w") as jfp:
    json.dump(jloco, jfp, indent=2)
    
