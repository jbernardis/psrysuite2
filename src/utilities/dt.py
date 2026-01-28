import os
from datetime import datetime
import openpyxl


class Breaker:
	def __init__(self, name):
		self.name = name
		self.state = 1
		self.tripTime = None
		self.tripDurations = []

	def UpdateState(self, tm, state):
		if state == self.state:
			return

		self.state = state
		if state == 0:
			self.tripTime = tm
		else:
			elapsed = tm - self.tripTime
			self.tripDurations.append(int(elapsed.total_seconds()))
			self.tripTime = None

	def TripDurations(self):
		return self.tripDurations


class Train:
	def __init__(self, iname, rname):
		self.iname = iname
		self.rname = rname
		self.stopped = False
		self.stopTime = None
		self.stopTimes = []
		self.blocks = []
		self.engineer = None
		self.engineers = []
		self.stopBlocks = []

	def Name(self):
		return self.iname if self.rname is None else self.rname

	def SetBlocks(self, blocks):
		self.blocks = blocks
		return blocks, []

	def UpdateBlocks(self, blocks):
		newBlocks = [b for b in blocks if b not in self.blocks]
		delBlocks = [b for b in self.blocks if b not in blocks]
		self.blocks = blocks
		return newBlocks, delBlocks

	def UpdateStoppage(self, tm, engineer, stop, block):
		if stop == self.stopped:
			return

		if stop:
			self.stopped = True
			self.stopTime = tm
			self.engineer = engineer
		else:
			self.stopped = False
			self.engineers.append(engineer)
			self.stopBlocks.append(block)
			elapsed = tm - self.stopTime
			self.stopTimes.append(int(elapsed.total_seconds()))
			self.stopTime = None
			self.engineer = None

	def StoppageTimes(self):
		return self.stopTimes, self.engineers, self.stopBlocks


class Block:
	def __init__(self, bn):
		self.name = bn
		self.entryTime = None
		self.traversalTimes = []
		self.trains = []
		self.engineers = []

	def EntryTime(self, tm):
		self.entryTime = tm

	def ExitTime(self, tm, trn, engineer):
		if self.entryTime is None:
			return
		elapsed = tm - self.entryTime
		self.traversalTimes.append(int(elapsed.total_seconds()))
		self.trains.append(trn)
		self.engineers.append(engineer)
		self.entryTime = None

	def TraversalTimes(self):
		return self.traversalTimes, self.trains, self.engineers


class LogReport:
	def __init__(self):
		self.blocks = {}
		self.trains = {}
		self.breakers = {}

	def ProcessLogFile(self, fname):
		# with open("../logs/rrserver.log", "r") as lfp:
		with open(fname, "r") as lfp:
			for line in lfp:
				ts = line[:23]
				dto = datetime.strptime(ts,"%Y-%m-%d %H:%M:%S,%f")
				rest = line[24:]
				if rest.startswith("sending message: "):

					j = eval("%s" % rest[17:])

					for cmd in j.keys():
						if cmd == "train":
							for p in j["train"]:
								self.ProcessTrainCommand(dto, p)

						elif cmd == ("breaker"):
							for p in j["breaker"]:
								self.ProcessBreakerCommand(dto, p)

	def ProcessTrainCommand(self, tm, parms):
		iname = parms["iname"]
		rname = parms["rname"]
		blks = parms["blocks"]
		stopped = parms["stopped"]
		engineer = parms["engineer"]

		if iname not in self.trains:
			tr = Train(iname, rname)
			self.trains[iname] = tr
			newBlks, delBlks = self.trains[iname].SetBlocks(blks)
		else:
			tr = self.trains[iname]
			newBlks, delBlks = tr.UpdateBlocks(blks)

		for b in newBlks:
			# print("Block %s entered at time %s" % (b, tm))
			if b not in self.blocks:
				blk = Block(b)
				self.blocks[b] = blk
			else:
				blk = self.blocks[b]

			blk.EntryTime(tm)

		for b in delBlks:
			if b in self.blocks:
				# print("Block %s exited at time %s" % (b, tm))
				blk = self.blocks[b]
				blk.ExitTime(tm, tr.Name(), engineer)
			else:
				# print("no record of block %s being entered" % b)
				pass

		tr.UpdateStoppage(tm, engineer, stopped, blks[0])  # TODO:  <== verify 0 or -1 here

	def ProcessBreakerCommand(self, tm, parms):
		name = parms["name"]
		value = parms["value"]
		if name in self.breakers:
			brk = self.breakers[name]
		else:
			brk = Breaker(name)
			self.breakers[name] = brk

		brk.UpdateState(tm, value)

	def ReportBlockTraversalTimes(self, worksheet):
		row = 2

		worksheet.title = "Block Traversal Times"
		worksheet.cell(row=1, column=1, value="Block")
		worksheet.cell(row=1, column=2, value="Seconds")
		worksheet.cell(row=1, column=3, value="Train")
		worksheet.cell(row=1, column=4, value="Engineer")

		for bn in sorted(self.blocks.keys()):
			blk = self.blocks[bn]
			trvTimes, trains, engineers = blk.TraversalTimes()
			for i in range(len(trvTimes)):
				tm = trvTimes[i]
				trn = trains[i]
				eng = engineers[i]
				worksheet.cell(row=row, column=1, value=bn)
				worksheet.cell(row=row, column=2, value=tm)
				worksheet.cell(row=row, column=3, value=trn)
				worksheet.cell(row=row, column=4, value=eng)
				row += 1

	def ReportTrainStoppageTimes(self, worksheet):
		row = 2

		worksheet.title = "Stopping Sections"
		worksheet.cell(row=1, column=1, value="Train")
		worksheet.cell(row=1, column=2, value="Block")
		worksheet.cell(row=1, column=3, value="Seconds")
		worksheet.cell(row=1, column=4, value="Engineer")

		for b in sorted(self.trains.keys()):
			trn = self.trains[b]
			tms, engs, blks = trn.StoppageTimes()
			if len(tms) > 0:
				for i in range(len(tms)):
					worksheet.cell(row=row, column=1, value=trn.Name())
					worksheet.cell(row=row, column=2, value=blks[i])
					worksheet.cell(row=row, column=3, value=tms[i])
					worksheet.cell(row=row, column=4, value=engs[i])
					row += 1

	def ReportBreakers(self, worksheet):
		row = 2

		worksheet.title = "Breakers"
		worksheet.cell(row=1, column=1, value="Name")
		worksheet.cell(row=1, column=2, value="Seconds")

		for b in sorted(self.breakers.keys()):
			brk = self.breakers[b]
			times = brk.TripDurations()
			if len(times) > 0:
				for tm in times:
					worksheet.cell(row=row, column=1, value=b)
					worksheet.cell(row=row, column=2, value=tm)
					row += 1


class ReportMain:
	def __init__(self):
		self.lr = LogReport()
		self.lr.ProcessLogFile("../logs/rrserver.log")

		wb = openpyxl.Workbook()
		ws = wb.active
		# ws.column_dimensions['B'].width = 10
		# ws.column_dimensions['C'].width = 10
		# ws.column_dimensions['D'].width = 50
		# ws.row_dimensions[1].height = 30

		self.lr.ReportBlockTraversalTimes(ws)

		wsstop = wb.create_sheet()
		self.lr.ReportTrainStoppageTimes(wsstop)

		wsbreakers = wb.create_sheet()
		self.lr.ReportBreakers(wsbreakers)

		xlsfn = os.path.join(os.getcwd(), "stats.xlsx")
		wb.save(xlsfn)

ReportMain()
