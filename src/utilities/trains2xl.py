import json
import os
import openpyxl
from openpyxl.styles import Alignment

class MainClass:
    def __init__(self):
        fn = os.path.join(os.getcwd(), "data", "newtrains.json")
        try:
            with open(fn, "r") as jfp:
                trains = json.load(jfp)
        except:
            print("Unable to open trains file: %s" % fn)
            exit(1)
        
        self.wb = openpyxl.Workbook()
        ws = self.wb.active 
        ws.title = "All Trains" 
        
        ws.cell(row=1, column=1, value="Train")
        ws.cell(row=1, column=2, value="Eastbound")
        ws.cell(row=1, column=3, value="Cutoff")
        ws.cell(row=1, column=4, value="Description")
        ws.cell(row=1, column=5, value="Normal Loco")
        ws.cell(row=1, column=6, value="Origin Loc")
        ws.cell(row=1, column=7, value="Origin Track")
        ws.cell(row=1, column=8, value="Terminus Loc")
        ws.cell(row=1, column=9, value="Terminus Track")
        ws.cell(row=1, column=10, value="Start Block")
        ws.cell(row=1, column=11, value="Start Sub Block")
        ws.cell(row=1, column=12, value="Time")
        
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        ws.row_dimensions[1].height = 30
        
        
 
        returnLink = "=HYPERLINK(\"#'All Trains'!A1\", \"Return\")"       
        row = 2
        for tr, tri in trains.items():
            ws.cell(row=row, column=1, value=tr)
            ws.cell(row=row, column=2, value=tri["eastbound"])
            ws.cell(row=row, column=3, value=tri["cutoff"])
            ws.cell(row=row, column=4, value=tri["desc"])
            ws.cell(row=row, column=5, value=tri["normalloco"])
            ws.cell(row=row, column=6, value=tri["origin"]["loc"])
            ws.cell(row=row, column=7, value=tri["origin"]["track"])
            ws.cell(row=row, column=8, value=tri["terminus"]["loc"])
            ws.cell(row=row, column=9, value=tri["terminus"]["track"])
            ws.cell(row=row, column=10, value=tri["startblock"])
            ws.cell(row=row, column=11, value=tri["startsubblock"])
            ws.cell(row=row, column=12, value=tri["time"])
            
            wstrk = self.CreateTrackerSheet(tr, tri, returnLink)
            wsseq = self.CreateSequenceSheet(tr, tri, returnLink)
            ws.cell(row=row, column=13, value="=HYPERLINK(\"#'" + wstrk.title + "'!A1\", \"TRK\")")
            ws.cell(row=row, column=14, value="=HYPERLINK(\"#'" + wsseq.title + "'!A1\", \"SEQ\")")
            row += 1
            
        ntrains = row-2
        #cell.value = '=HYPERLINK("#Sheet2!A2", "click here")'
        
        row = ws[1] # header row
        for c in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
            cell = row[c]
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        for row in ws[2:ws.max_row]:  # skip the header
            for c in [0, 1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
                cell = row[c]
                cell.alignment = Alignment(horizontal='center')        
        
            
        xlsfn = os.path.join(os.getcwd(), "data", "trains.xlsx") 
        self.wb.save(xlsfn)
        
        print("%d trains saved to spreadsheet: %s" % (ntrains, xlsfn))
        
    def CreateTrackerSheet(self, tr, tri, link):
        wstrk = self.wb.create_sheet("%s tracker" % tr)
        wstrk.cell(row=1, column=4, value=link)
        row = 2
        for step in tri["tracker"]:
            wstrk.cell(row=row, column=1, value=step[0])
            wstrk.cell(row=row, column=2, value=step[1])
            wstrk.cell(row=row, column=3, value=step[2])
            row += 1
            
        wstrk.column_dimensions['A'].width = 14
        wstrk.column_dimensions['B'].width = 40
        wstrk.row_dimensions[1].height = 30
           
        for row in wstrk[2:wstrk.max_row]:  # skip the header
            cell = row[2]
            cell.alignment = Alignment(horizontal='center')        
        
        row = wstrk[1] # header row
        for c in [0, 1, 2, 3]:
            cell = row[c]
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
        return wstrk
    
    def CreateSequenceSheet(self, tr, tri, link):
        wsseq = self.wb.create_sheet("%s sequence" % tr)
        wsseq.cell(row=1, column=7, value=link)
        wsseq.cell(row=1, column=1, value="block")
        wsseq.cell(row=1, column=2, value="os")
        wsseq.cell(row=1, column=3, value="route")
        wsseq.cell(row=1, column=4, value="signal")
        wsseq.cell(row=1, column=5, value="time")
        wsseq.cell(row=1, column=6, value="trigger")
        row = 2
        for step in tri["sequence"]:
            wsseq.cell(row=row, column=1, value=step["block"])
            wsseq.cell(row=row, column=2, value=step["os"])
            wsseq.cell(row=row, column=3, value=step["route"])
            wsseq.cell(row=row, column=4, value=step["signal"])
            wsseq.cell(row=row, column=5, value=step["time"])
            wsseq.cell(row=row, column=6, value=step["trigger"])
            row += 1

        wsseq.column_dimensions['A'].width = 14
        wsseq.column_dimensions['B'].width = 14
        wsseq.column_dimensions['C'].width = 14
        wsseq.column_dimensions['D'].width = 14
        wsseq.column_dimensions['E'].width = 14
        wsseq.column_dimensions['F'].width = 14
        wsseq.row_dimensions[1].height = 30
        
        for row in wsseq[2:wsseq.max_row]:  # skip the header
            for c in [0, 1, 2, 3, 4, 5]:
                cell = row[c]
                cell.alignment = Alignment(horizontal='center')        
        
        row = wsseq[1] # header row
        for c in [0, 1, 2, 3, 4, 5, 6]:
            cell = row[c]
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
           
        return wsseq


        
mc = MainClass()
        