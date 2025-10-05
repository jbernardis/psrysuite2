import openpyxl
import os
import json
from pip._internal.operations.build import build_tracker

class MainUnit:
    def __init__(self):
        fn = os.path.join(os.getcwd(), "data", "trains.xlsx")
        self.wb = openpyxl.load_workbook(filename = fn)
        ws = self.wb.active
        
        row = 2  # skip title row
        
        jtrains = {}
        while True:
            train = ws.cell(row=row, column=1).value
            if train is None:
                break
           
            eastbound = ws.cell(row=row, column=2).value 
            cutoff = ws.cell(row=row, column=3).value
            desc = ws.cell(row=row, column=4).value
            normalloco = ws.cell(row=row, column=5).value
            originloc = ws.cell(row=row, column=6).value
            origintrack = ws.cell(row=row, column=7).value
            terminusloc = ws.cell(row=row, column=8).value
            terminustrack = ws.cell(row=row, column=9).value
            startblock = ws.cell(row=row, column=10).value
            startsubblock = ws.cell(row=row, column=11).value
            time = ws.cell(row=row, column=12).value
            
            jtrains[train] = {
                "eastbound": eastbound, 
                "cutoff": cutoff,
                "desc": desc,
                "block": None,
                "loco": normalloco,
                "normalloco": normalloco,
                "origin": {
                    "loc": originloc,
                    "track": origintrack
                    },
                "terminus": {
                    "loc": terminusloc,
                    "track": terminustrack
                    },
                "startblock": startblock,
                "startsubblock": startsubblock,
                "time": time,
                "tracker": self.GetTrackerWorksheet(train),
                "sequence": self.GetSequenceWorksheet(train)
                }
            
            row += 1
            
        jfn = os.path.join(os.getcwd(), "data", "newnewtrains.json")
        with open(jfn, "w") as jfp:
            json.dump(jtrains, jfp, indent=2)

            
    def GetTrackerWorksheet(self, train):
        sheetName = "%s tracker" % train
        ws = self.wb[sheetName]
        row = 2  # skip title row
        
        tracker = []
        while True:
            v1 = ws.cell(row=row, column=1).value
            v2 = ws.cell(row=row, column=2).value
            v3 = ws.cell(row=row, column=3).value
            if v1 is None and v2 is None and v3 is None:
                break
            
            tracker.append([v1, v2, v3])
            row += 1
            
        return tracker
    
    def GetSequenceWorksheet(self, train):
        sheetName = "%s sequence" % train
        ws = self.wb[sheetName]
        row = 2  # skip title row
        
        sequence = []
        while True:
            block = ws.cell(row=row, column=1).value
            if block is None:
                break
            osblk = ws.cell(row=row, column=2).value
            route = ws.cell(row=row, column=3).value
            signal = ws.cell(row=row, column=4).value
            time = ws.cell(row=row, column=5).value
            trigger = ws.cell(row=row, column=6).value
            
            seq = {
                "block": block,
                "os": osblk,
                "route": route,
                "signal": signal,
                "time": time,
                "trigger": trigger
                }
             
            sequence.append(seq)
            row += 1
            
        return sequence
    
m = MainUnit()
            
