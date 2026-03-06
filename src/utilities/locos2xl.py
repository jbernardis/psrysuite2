import json
import os
import openpyxl

fdir = os.path.abspath(os.path.join(os.getcwd(), '..'))
print("fdir = %s" % fdir)
fn = os.path.join(fdir, "data", "locos.json")
print("fn = %s" % fn)
try:
    with open(fn, "r") as jfp:
        locos = json.load(jfp)
except:
    print("Unable to open locos file: %s" % fn)
    exit(1)

wb = openpyxl.Workbook()
ws = wb.active  

ws.cell(row=1, column=1, value="Loco")
ws.cell(row=1, column=2, value="Short")
ws.cell(row=1, column=3, value="Description")
ws.cell(row=1, column=4, value="Start")
ws.cell(row=1, column=5, value="Slow")
ws.cell(row=1, column=6, value="Medium")
ws.cell(row=1, column=7, value="Fast")
ws.cell(row=1, column=8, value="Acc")
ws.cell(row=1, column=9, value="Dec")

row = 2
for l, li in locos.items():
    ws.cell(row=row, column=1, value=l)
    ws.cell(row=row, column=2, value="True" if li["short"] else "False")
    ws.cell(row=row, column=3, value=li["desc"])
    ws.cell(row=row, column=4, value=li["prof"]["start"])
    ws.cell(row=row, column=5, value=li["prof"]["slow"])
    ws.cell(row=row, column=6, value=li["prof"]["medium"])
    ws.cell(row=row, column=7, value=li["prof"]["fast"])
    ws.cell(row=row, column=8, value=li["prof"]["acc"])
    ws.cell(row=row, column=9, value=li["prof"]["dec"])
    row += 1
    
xlsfn = os.path.join(fdir, "data", "locos.xlsx")
wb.save(xlsfn)

print("%d locomotives saved to spreadsheet: %s" % (row-2, xlsfn))
    