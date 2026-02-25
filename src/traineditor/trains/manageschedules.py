import wx
import os
import subprocess
import qrcode
import logging
import openpyxl
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill

import utilities.HTML as HTML
from traineditor.reports import Report
from traineditor.trains.choosetrains import ChooseTrainsDlg, ChooseScheduleDlg
from traineditor.trains.schedule import Schedule

BTNSZ = (120, 46)


class ManageSchedules:
	def __init__(self, parent, roster, locos, trainList, rrserver, settings):
		self.roster = roster
		self.locos = locos
		self.trains = trainList
		self.parent = parent
		self.rrserver = rrserver
		self.settings = settings
		self.browser = self.settings.browser
		self.spreadsheet = self.settings.spreadsheet

		dlg = ChooseTrainsDlg(self.parent, trainList, rrserver, self.TrainCardsReport, self.ScheduleReport)
		rc = dlg.ShowModal()

		if rc == wx.ID_OK:
			sch = dlg.getResults()

		dlg.Destroy()

		if rc != wx.ID_OK:
			return

	def TrainCardsReport(self, sched):
		r = TrainCardsReport(self.parent, self.browser)
		r.TrainCards(self.roster, sched)

	def ScheduleReport(self, sched):
		dlg = SchedParmsDlg(self.parent)
		rc = dlg.ShowModal()
		if rc != wx.ID_OK:
			dlg.Destroy()
			return

		live, html, excel = dlg.GetResults()
		dlg.Destroy()

		r = SchedulesReport(self.parent, self.browser, self.spreadsheet, live)

		if html:
			r.ScheduleReportHTML(self.roster, self.locos, sched, self.rrserver)

		if excel:
			r.ScheduleReportXL(self.roster, self.locos, sched, self.rrserver)


class SchedulesReport(Report):
	def __init__(self, parent, browser, spreadsheet, live):
		Report.__init__(self, parent, browser, spreadsheet)
		self.parent = parent
		self.roster = None
		self.locos = None
		self.useLiveData = live

		self.passengerFill = PatternFill(start_color="FF8AFFBB", fill_type="solid")
		self.freightFill = PatternFill(start_color="FFFFB6B2", fill_type="solid")

	def ScheduleReportHTML(self, roster, locos, sched, rrserver):
		self.roster = roster
		self.locos = locos

		if sched is None:
			dlg = wx.MessageDialog(self.parent, "No Schedule Information provided",
				"No Schedule Specified",
				wx.OK | wx.ICON_INFORMATION)

			dlg.ShowModal()
			dlg.Destroy()
			return

		css = HTML.StyleSheet()
		css.addElement("table", {"border-collapse": "collapse", "border-spacing": "0", "width": "170mm",
								 "font-family": 'Arial, sans-serif', "font-size": "20px", "margin-left": "auto",
								 "margin-right": "auto"})
		css.addElement("th", {'text-align': 'center', 'overflow': 'hidden'})
		css.addElement("td", {'text-align': 'center', 'overflow': 'hidden'})
		css.addElement("table, th, td", {'border': "1px solid black", 'border-collapse': 'collapse'})
		css.addElement("td.index", {"width": "10mm", "font-weight": "bold"})
		css.addElement("td.trainid", {"width": "20mm", "font-weight": "bold"})
		css.addElement("td.loco", {"width": "20mm", "font-weight": "bold"})
		css.addElement("td.dir", {"width": "10mm", "font-weight": "bold"})
		css.addElement("td.engineer", {"width": "60mm"})
		css.addElement("td.origin", {"width": "25mm", "font-weight": "bold"})
		css.addElement("td.terminus", {"width": "25mm", "font-weight": "bold"})
		css.addElement("td.freight", {"background-color": "#FFB6B2"})
		css.addElement("td.passenger", {"background-color": "#8AFFBB"})
		css.addElement("p.header", {"font-family": 'Arial, sans-serif', "font-size": "30px", "font-weight": "bold"})

		html = HTML.starthtml()
		html += HTML.head(HTML.style({'type': "text/css", 'media': "screen, print"}, css))

		html += HTML.startbody()

		header = HTML.tr({},
					HTML.th({}, ""),
					HTML.th({}, "Train"),
					HTML.th({}, "Loco"),
					HTML.th({}, "Dir"),
					HTML.th({}, "Engineer"),
					HTML.th({}, "Origin"),
					HTML.th({}, "Terminus"),
					)

		activetrains = rrserver.Get("activetrains", {})
		logging.debug("active trains = %s" % str(activetrains))

		rows = []
		rowx = 1
		for tid in sched.getSchedule():
			tinfo = activetrains.get(tid, None) if self.useLiveData else None
			rows.append(self.formatTableRow(tid, tinfo, rowx))
			rowx += 1

		if len(rows) > 0:
			html += HTML.div({"style": "height: 50px"})
			html += HTML.p({'align': 'center', 'class': 'header'}, "Main Schedule")
			html += HTML.table({}, header, "".join(rows))

		rows = []
		rowx = 1
		for tid in sched.getExtras():
			tinfo = activetrains.get(tid, None) if self.useLiveData else None
			rows.append(self.formatTableRow(tid, tinfo, rowx, alpha=True))
			rowx += 1

		if len(rows) > 0:
			html += HTML.div({"style": "height: 50px"})
			html += HTML.p({'align': 'center', 'class': 'header'}, "Extra Trains")
			html += HTML.table({}, header, "".join(rows))

		html += HTML.endbody()
		html += HTML.endhtml()

		self.openBrowser("Schedule", html)

	def formatTableRow(self, tid, tinfo, rowx, alpha=False):
		r = self.roster.get(tid, None)
		if r is None:
			east = True
			origin = ""
			terminus = ""
			loco = None
		else:
			east = r["eastbound"]
			origin = "%s" % r["origin"]["loc"]
			trk = r["origin"]["track"]
			if tinfo is not None:
				blks = tinfo.get("blocks", None)
				if blks is not None and len(blks) > 0:
					trk = blks[0]
			if trk is not None:
				origin += "(%s)" % trk
			terminus = "%s" % r["terminus"]["loc"]
			trk = r["terminus"]["track"]
			if trk is not None:
				terminus += "(%s)" % trk
			loco = r.get("normalloco", None)

		if tinfo is not None:
			aloco = tinfo.get("loco", None)
			if aloco is not None and aloco != "??":
				loco = aloco

		if loco is None:
			loco = ""
		else:
			linfo = self.locos.getLoco(loco)
			if linfo is not None and linfo["short"]:
				loco += "(s)"

		passenger = tid[0].isdigit()

		engineer = ""

		colorClass = "passenger" if passenger else "freight"

		if alpha:
			index = chr(ord('A') + rowx - 1)
		else:
			index = "%-2d" % rowx

		return HTML.tr({},
				"\n   " + HTML.td({"class": "index"}, index) + "\n",
				"  " + HTML.td({"class": "trainid %s" % colorClass}, tid) + "\n",
				"  " + HTML.td({"class": "loco"}, loco) + "\n",
				"  " + HTML.td({"class": "dir"}, "E" if east else "W") + "\n",
				"  " + HTML.td({"class": "engineer"}, engineer) + "\n",
				"  " + HTML.td({"class": "origin"}, origin) + "\n",
				"  " + HTML.td({"class": "terminus"}, terminus) + "\n"
				) + "\n"

	def ScheduleReportXL(self, roster, locos, sched, rrserver):
		self.roster = roster
		self.locos = locos

		if sched is None:
			dlg = wx.MessageDialog(self.parent, "No Schedule Information provided",
				"No Schedule Specified",
				wx.OK | wx.ICON_INFORMATION)

			dlg.ShowModal()
			dlg.Destroy()
			return

		activetrains = rrserver.Get("activetrains", {})

		wb = openpyxl.Workbook()
		ws = wb.active

		ws.column_dimensions['A'].width = 6
		ws.column_dimensions['B'].width = 10
		ws.column_dimensions['C'].width = 10
		ws.column_dimensions['D'].width = 6
		ws.column_dimensions['E'].width = 30
		ws.column_dimensions['F'].width = 14
		ws.column_dimensions['G'].width = 14

		ft14bold = Font(name='Arial', bold=True, size=14)
		ft14 = Font(name='Arial', size=14)
		aligncenter = Alignment(horizontal="center")

		ws.merge_cells('A2:G2')
		top_left_cell = ws['A2']
		top_left_cell.value = "Main Schedule"
		top_left_cell.alignment = Alignment(horizontal="center")
		top_left_cell.font = Font(name='Arial', bold=True, size=24)

		ws.title = "Schedule"
		ws.cell(row=4, column=1, value="")
		ws.cell(row=4, column=2, value="Train")
		ws.cell(row=4, column=3, value="Loco")
		ws.cell(row=4, column=4, value="Dir")
		ws.cell(row=4, column=5, value="Engineer")
		ws.cell(row=4, column=6, value="Origin")
		ws.cell(row=4, column=7, value="Terminus")

		rowx = 5
		index = 1
		for tid in sched.getSchedule():
			tinfo = activetrains.get(tid, None) if self.useLiveData else None
			self.formatSheetRow(tid, tinfo, ws, rowx, index)
			rowx += 1
			index += 1

		border = Border(
			left=Side(border_style="thin", color='00000000'),
			right=Side(border_style="thin", color='00000000'),
			top=Side(border_style="thin", color='FF000000'),
			bottom=Side(border_style="thin", color='FF000000'))

		for row in ws["A4:G4"]:
			for cell in row:
				cell.font = ft14bold
				cell.alignment = aligncenter
				cell.border = border

		for row in ws["A5:G%s" % (rowx-1)]:
			for cell in row:
				cell.font = ft14
				cell.alignment = aligncenter
				cell.border = border

		if sched.lenExtras() > 0:
			rowx += 2

			ws.merge_cells("A%d:G%d" % (rowx, rowx))
			top_left_cell = ws['A%d' % rowx]
			top_left_cell.value = "Extra Trains"
			top_left_cell.alignment = Alignment(horizontal="center")
			top_left_cell.font = Font(name='Arial', bold=True, size=24)

			rowx += 2
			index = 1
			startrow = rowx
			for tid in sched.getExtras():
				tinfo = activetrains.get(tid, None) if self.useLiveData else None
				self.formatSheetRow(tid, tinfo, ws, rowx, index, alpha=True)
				rowx += 1
				index += 1

			for row in ws["A%d:G%s" % (startrow, (rowx-1))]:
				for cell in row:
					cell.font = ft14
					cell.alignment = aligncenter
					cell.border = border

		xlsfn = os.path.join(os.getcwd(), "report.xlsx")
		wb.save(xlsfn)

		if self.spreadsheet is not None:
			process = subprocess.Popen([self.spreadsheet, xlsfn])

	def formatSheetRow(self, tid, tinfo, ws, rowx, index, alpha=False):
		r = self.roster.get(tid, None)
		if r is None:
			east = True
			origin = ""
			terminus = ""
			loco = None
		else:
			east = r["eastbound"]
			origin = "%s" % r["origin"]["loc"]
			trk = r["origin"]["track"]
			if tinfo is not None:
				blks = tinfo.get("blocks", None)
				if blks is not None and len(blks) > 0:
					trk = blks[0]
			if trk is not None:
				origin += "(%s)" % trk
			terminus = "%s" % r["terminus"]["loc"]
			trk = r["terminus"]["track"]
			if trk is not None:
				terminus += "(%s)" % trk
			loco = r.get("normalloco", None)

		if tinfo is not None:
			aloco = tinfo.get("loco", None)
			if aloco is not None:
				loco = aloco

		if loco is None:
			loco = ""
		else:
			linfo = self.locos.getLoco(loco)
			if linfo is not None and linfo["short"]:
				loco += "(s)"

		passenger = tid[0].isdigit()

		engineer = ""

		colorClass = self.passengerFill if passenger else self.freightFill

		if alpha:
			idx = chr(ord('A') + index - 1)
		else:
			idx = "%-2d" % index

		ws.cell(row=rowx, column=1, value=idx)
		ws.cell(row=rowx, column=2, value=tid)
		ws.cell(row=rowx, column=2).fill = colorClass
		ws.cell(row=rowx, column=3, value=loco)
		ws.cell(row=rowx, column=4, value="E" if east else "W")
		ws.cell(row=rowx, column=5, value=engineer)
		ws.cell(row=rowx, column=6, value=origin)
		ws.cell(row=rowx, column=7, value=terminus)


class TrainCardsReport (Report):
	def __init__(self, parent, browser):
		Report.__init__(self, parent, browser, None)
		self.roster = None

	def TrainCards(self, roster, sched):
		self.roster = roster

		if sched is None:
			dlg = wx.MessageDialog(self.parent, "No Schedule Information provided",
				"No Schedule Specified",
				wx.OK | wx.ICON_INFORMATION)

			dlg.ShowModal()
			dlg.Destroy()
			return

		ct = 0
		for flag in sched.getSchedule():
			if flag:
				ct += 1

		ctx = 0
		for flag in sched.getExtras():
			if flag:
				ctx += 1

		if ct+ctx == 0:
			dlg = wx.MessageDialog(self.parent, "No Train Cards chosen - skipping report",
						"Nothing to print",
						wx.OK | wx.ICON_INFORMATION)
			dlg.ShowModal()
			dlg.Destroy()
			return

		css = HTML.StyleSheet()
		css.addElement("div.page", {"page-break-inside": "avoid"})
		css.addElement("*", {"box-sizing": "border-box"})
		css.addElement(".row", {"margin-left": "-5px", "margin-right": "-5px", "height": "107mm"})
		css.addElement(".column", {"float": "left", "width": "136mm", "padding": "1px"})
		css.addElement(".row::after", {"content": '""', "clear": "both", "display": "table"})
		css.addElement("tr.qrrow", {"height": "15mm"})
		css.addElement("table", {"border-collapse": "collapse", "border-spacing": "0", "width": "100%", "height": "106mm", "font-family": '"Times New Roman", Times, serif', "font-size": "16px"})
		css.addElement("td.trainid", {"width": "36.4%", "padding-left": "50px", "padding-top": "10px", "font-size": "28px", "font-weight": "bold"})
		css.addElement("td.qr", {"width": "36.4%", "padding-left": "150px"})
		css.addElement("td.firstcol", {"width": "36.4%", "padding-left": "50px"})
		css.addElement("td.secondcol", {"width": "10%"})
		css.addElement("tr.datarow", {"height": "5mm"})
		css.addElement("tr.descrow", {"height": "5mm"})
		css.addElement("td", {"text-align": "left", "padding-left": "6px"})
		css.addElement("td.cardnumber", {"text-align": "right", "padding-right": "50px"})

		html  = HTML.starthtml()
		html += HTML.head(HTML.style({'type': "text/css", 'media': "screen, print"}, css))

		html += HTML.startbody()

		cards = []

		tx = 0
		for tid in sched.getSchedule():
			self.TrainQRCode(tid)
			cards.append(self.formatTrainCard(tid, roster[tid], "%d" % (tx+1)))
			tx += 1

		tx = 0
		for tid in sched.getExtras():
			self.TrainQRCode(tid)
			cn = chr(ord('A') + tx)
			cards.append(self.formatTrainCard(tid, roster[tid], cn))
			tx += 1

		nCards = len(cards)

		divs = []
		for i in range(0, nCards-1, 2):
			divs.append(HTML.div({"class": "row"}, cards[i], cards[i+1]))

		if nCards%2 != 0:
			divs.append(HTML.div({"class": "row"}, cards[-1]))

		dx = 0
		while dx < len(divs):
			if dx == len(divs)-1:
				html += HTML.div({"class": "page"}, divs[dx])
			else:
				html += HTML.div({"class": "page"}, divs[dx], divs[dx+1])

			dx += 2

		html += HTML.endbody()
		html += HTML.endhtml()

		self.openBrowser("Train Cards", html)

	def TrainQRCode(self, tid):
		qr = qrcode.QRCode(
			version=1,
			error_correction=qrcode.constants.ERROR_CORRECT_L,
			box_size=3,
			border=4,
		)
		qr.add_data("TRAIN: %s" % tid)
		qr.make(fit=True)
		img = qr.make_image(fill_color="black", back_color="white")
		# img = qrcode.make('TRAIN: CFYD')
		type(img)  # qrcode.image.pil.PilImage
		fn = os.path.join(os.getcwd(), "qrcodes", "train_%s.png" % tid)
		img.save(fn)

	def formatTrainCard(self, tid, tinfo, tx):
		fn = os.path.join("qrcodes", "train_%s.png" % tid)
		img = HTML.img({"src": fn})
		trainIdRow = HTML.tr({}, HTML.td({"class": "trainid", "colspan": "2"}, tid), HTML.td({"class": "qr"}, img))
		emptyRow = HTML.tr({"class": "datarow"}, HTML.td({}, HTML.nbsp()))
		descr = "%sbound %s" % ("East" if tinfo["eastbound"] else "West", tinfo["desc"])
		if tinfo["cutoff"]:
			descr += " (via cutoff)"
		descRow = HTML.tr({"class": "descrow"}, HTML.td({"class": "firstcol", "colspan": "3"}, descr))
		cardNumberRow = HTML.tr({}, HTML.td({}, ""), HTML.td({}, ""), HTML.td({"class": "cardnumber"}, tx))

		stepRows = []
		for stp in tinfo["tracker"]:
			row = HTML.tr({"class": "datarow"},
						HTML.td({"class": "firstcol"}, stp[0]),
						HTML.td({"class": "secondcol"}, "" if stp[2] == 0 else ("(%2d)" % stp[2])),
						HTML.td({}, stp[1])
			)
			stepRows.append(row)

		nRows = len(stepRows)
		nEmpty = 10 - nRows

		table = HTML.table({},
			trainIdRow,
			descRow,
			emptyRow,
			" ".join(stepRows),
			nEmpty * emptyRow,
			cardNumberRow
		)

		return HTML.div({"class": "column"}, table)


class SchedParmsDlg(wx.Dialog):
	def __init__(self, parent):
		wx.Dialog.__init__(self, parent, style=wx.DEFAULT_FRAME_STYLE)
		self.Bind(wx.EVT_CLOSE, self.OnClose)

		self.cbLiveData = wx.CheckBox(self, wx.ID_ANY, "Use Live Data")
		self.cbLiveData.SetValue(True)

		self.cbHTML = wx.CheckBox(self, wx.ID_ANY, "Produce HTML report")
		self.cbHTML.SetValue(True)

		self.cbExcel = wx.CheckBox(self, wx.ID_ANY, "Produce Excel report")
		self.cbExcel.SetValue(False)

		self.bOK = wx.Button(self, wx.ID_ANY, "OK")
		self.Bind(wx.EVT_BUTTON, self.OnBOK, self.bOK)

		self.bCancel = wx.Button(self, wx.ID_ANY, "Cancel")
		self.Bind(wx.EVT_BUTTON, self.OnBCancel, self.bCancel)

		vsz = wx.BoxSizer(wx.VERTICAL)
		vsz.AddSpacer(20)

		vsz.Add(self.cbLiveData, 0, wx.LEFT, 40)
		vsz.AddSpacer(20)

		vsz.Add(self.cbHTML, 0, wx.LEFT, 40)
		vsz.AddSpacer(5)

		vsz.Add(self.cbExcel, 0, wx.LEFT, 40)
		vsz.AddSpacer(30)

		bsz = wx.BoxSizer(wx.HORIZONTAL)
		bsz.AddSpacer(20)
		bsz.Add(self.bOK)
		bsz.AddSpacer(30)
		bsz.Add(self.bCancel)
		bsz.AddSpacer(20)

		vsz.Add(bsz)
		vsz.AddSpacer(20)

		hsz = wx.BoxSizer(wx.HORIZONTAL)
		hsz.AddSpacer(20)
		hsz.Add(vsz)
		hsz.AddSpacer(20)

		self.SetSizer(hsz)
		self.Fit()
		self.Layout()

	def OnClose(self, _):
		self.EndModal(wx.ID_CANCEL)

	def OnBCancel(self, _):
		self.EndModal(wx.ID_CANCEL)

	def OnBOK(self, _):
		self.EndModal(wx.ID_OK)

	def GetResults(self):
		return self.cbLiveData.IsChecked(), self.cbHTML.IsChecked(), self.cbExcel.IsChecked()




